import os
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# 忽略 UserWarning 类型的警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def define_default_font():
    """
    定义一个默认的字体样式

    Returns:
        Font: 默认字体样式对象
    """
    return Font(name='Arial', size=11, bold=False, italic=False,
                vertAlign=None, underline='none', strike=False,
                color='FF000000')


def apply_default_font_to_workbook(workbook):
    """
    应用默认样式到所有工作表的所有单元格

    Args:
        workbook (Workbook): 要应用样式的Excel工作簿对象
    """
    default_font = define_default_font()
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = default_font


def create_template_workbook():
    """
    创建一个具有默认样式的模板工作簿

    Returns:
        Workbook: 具有默认样式的Excel模板工作簿对象
    """
    template_wb = Workbook()
    apply_default_font_to_workbook(template_wb)
    return template_wb


def combine_excel_files(directory, output_file):
    """
    合并指定文件夹下的所有Excel文件到一个新的工作簿

    Args:
        directory (str): 包含Excel文件的文件夹路径
        output_file (str): 合并后输出的Excel文件名
    """
    # 创建一个新的工作簿用于存放合并的数据
    merged_wb = create_template_workbook()
    merged_ws = merged_wb.active

    # 初始化列标题行
    headers_added = False

    # 遍历指定目录下的所有文件
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                # 构建完整的文件路径
                filepath = os.path.join(directory, filename)

                # 使用openpyxl加载Excel文件
                wb = load_workbook(filepath)

                # 应用默认样式
                apply_default_font_to_workbook(wb)

                # 获取第一个工作表
                ws = wb.active

                # 获取第一行作为列标题
                if not headers_added:
                    header_row = [cell.value for cell in next(ws.iter_rows())]
                    merged_ws.append(header_row)
                    headers_added = True

                # 添加除第一行外的数据
                for row in ws.iter_rows(min_row=2, values_only=True):
                    merged_ws.append(row)
            except Exception as e:
                print(f"Failed to process {filename}: {e}")

    # 保存合并后的数据到新的Excel文件
    merged_wb.save(output_file)
    print(f"Combined data has been written to {output_file}")


# 指定需要读取的文件夹路径和输出文件名
directory = r"C:\Users\Administrator\Desktop\XXX"
output_file = 'Output.xlsx'

combine_excel_files(directory, output_file)
